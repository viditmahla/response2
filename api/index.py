from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel
from typing import Optional
import uuid
from datetime import datetime, timezone
import json

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

DATA_CATEGORIES = {
    "baseline":      {"label": "Baseline",       "description": "Standard calcite ERW baseline data"},
    "feedstock_3x3": {"label": "3×3 Feedstock",  "description": "Multi-feedstock grid dataset"},
    "omega":         {"label": "Omega Variants",  "description": "Omega threshold variation data"},
}

# ── Models ──
class ChatRequest(BaseModel):
    message: str
    session_id: Optional[str] = "default"

class ChatResponse(BaseModel):
    reply: str
    session_id: str


def safe_float(v):
    if v is None: return None
    try:
        f = float(v)
        return f if f == f else None
    except (ValueError, TypeError): return None

def safe_int(v):
    if v is None: return None
    try: return int(v)
    except (ValueError, TypeError): return None


def parse_erw_row(row, feedstock_name, omega_threshold, data_category):
    return {
        "id": str(uuid.uuid4()),
        "feedstock": feedstock_name.lower(),
        "omega_threshold": omega_threshold,
        "data_category": data_category,
        "sample_no":      str(row[0]) if row[0] else "",
        "river_type":     str(row[1]) if row[1] else "",
        "latitude":       safe_float(row[2]),
        "longitude":      safe_float(row[3]),
        "ph":             safe_float(row[4]),
        "alkalinity":     safe_float(row[5]),
        "temp_c":         safe_float(row[6]),
        "ca":             safe_float(row[7]),
        "mg":             safe_float(row[8]),
        "na":             safe_float(row[9]),
        "k":              safe_float(row[10]),
        "cl":             safe_float(row[11]),
        "so4":            safe_float(row[12]),
        "no3":            safe_float(row[13]),
        "salinity":       safe_float(row[15]),
        "ksp":            safe_float(row[17]),
        "hco3":           safe_float(row[20]),
        "co3":            safe_float(row[21]),
        "co2_aq":         safe_float(row[22]),
        "dic":            safe_float(row[23]),
        "pco2":           safe_float(row[24]),
        "fco2":           safe_float(row[25]),
        "z_plus":         safe_float(row[27]),
        "z_minus":        safe_float(row[28]),
        "nicb":           safe_float(row[29]),
        "omega_calcite":  safe_float(row[31]),
        "si_calcite":     safe_float(row[32]),
        "state":          str(row[34]) if row[34] else "",
        "region":         str(row[35]) if row[35] else "",
        "river_name":     str(row[36]) if row[36] else "",
        "discharge":      safe_float(row[37]),
        "source":         str(row[38]) if row[38] else "",
        "j_steps":        safe_int(row[40]),
        "k_steps":        safe_int(row[41]),
        "rock_addition":  safe_float(row[42]),
        "omega_flag":     safe_int(row[43]),
        "success_flag":   safe_int(row[44]),
        "omega_final":    safe_float(row[47]),
        "ca_final":       safe_float(row[48]),
        "alk_final":      safe_float(row[50]),
        "dic_final":      safe_float(row[51]),
        "ph_final":       safe_float(row[53]),
        "pco2_final":     safe_float(row[55]),
        "discharge_ms":   safe_float(row[57]),
        "cdr_mol_s":      safe_float(row[58]),
        "cdr_t_yr":       safe_float(row[59]),
        "cdr_kt_yr":      safe_float(row[60]),
    }


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()


# ── Upload Status ──
@api_router.get("/upload/status")
async def upload_status():
    """Return counts per data_category so the UI knows what's loaded."""
    result = {}
    for cat in DATA_CATEGORIES:
        result[cat] = await db.erw_samples.count_documents({"data_category": cat})
    result["total"] = await db.erw_samples.count_documents({})
    return result


# ── Upload ──
@api_router.post("/feedstock/upload")
async def upload_feedstock(
    file: UploadFile = File(...),
    feedstock_name: str = "unknown",
    omega_threshold: int = 5,
    data_category: str = "baseline",
):
    if data_category not in DATA_CATEGORIES:
        raise HTTPException(status_code=400, detail=f"data_category must be one of {list(DATA_CATEGORIES.keys())}")
    try:
        import openpyxl, io
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        # ── ERW Results sheet ──
        ws_name = "ERW Results" if "ERW Results" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[ws_name]
        samples = []
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            if not row[0] or str(row[0]).strip() == '':
                continue
            if row[2] is None and row[3] is None and row[4] is None:
                continue
            samples.append(parse_erw_row(row, feedstock_name, omega_threshold, data_category))

        if samples:
            await db.erw_samples.insert_many(samples)

        # ── Summary Statistics sheet (optional) ──
        summaries_count = 0
        if 'Summary Statistics' in wb.sheetnames:
            ws2 = wb['Summary Statistics']
            summaries = []
            for row in ws2.iter_rows(min_row=4, max_row=ws2.max_row, values_only=True):
                if not row[0] or str(row[0]).strip() == '':
                    continue
                summaries.append({
                    "id": str(uuid.uuid4()),
                    "feedstock": feedstock_name.lower(),
                    "omega_threshold": omega_threshold,
                    "data_category": data_category,
                    "region":       str(row[0]),
                    "add_mean":     safe_float(row[1]) or 0,
                    "add_median":   safe_float(row[2]) or 0,
                    "add_std":      safe_float(row[3]) or 0,
                    "add_min":      safe_float(row[4]) or 0,
                    "add_max":      safe_float(row[5]) or 0,
                    "n_samples":    safe_int(row[6]) or 0,
                    "omega_mean":   safe_float(row[7]) or 0,
                    "omega_median": safe_float(row[8]) or 0,
                    "omega_std":    safe_float(row[9]) or 0,
                    "cdr_mean":     safe_float(row[10]) or 0,
                    "cdr_total":    safe_float(row[11]) or 0,
                    "n_with_q":     safe_int(row[12]) or 0,
                    "success_pct":  safe_float(row[13]) or 0,
                })
            if summaries:
                await db.summary_stats.insert_many(summaries)
                summaries_count = len(summaries)

        # ── Feedstocks registry ──
        existing = await db.feedstocks.find_one({"name": feedstock_name.lower(), "data_category": data_category})
        if existing:
            thresholds = existing.get("omega_thresholds", [])
            if omega_threshold not in thresholds:
                thresholds.append(omega_threshold)
            await db.feedstocks.update_one(
                {"name": feedstock_name.lower(), "data_category": data_category},
                {"$set": {"omega_thresholds": thresholds,
                          "sample_count": existing.get("sample_count", 0) + len(samples)}}
            )
        else:
            await db.feedstocks.insert_one({
                "id": str(uuid.uuid4()),
                "name": feedstock_name.lower(),
                "data_category": data_category,
                "omega_thresholds": [omega_threshold],
                "created_at": datetime.now(timezone.utc).isoformat(),
                "sample_count": len(samples),
            })

        return {
            "message": f"Uploaded {len(samples)} samples ({summaries_count} summary rows)",
            "samples_count": len(samples),
            "data_category": data_category,
        }
    except Exception as e:
        logger.error(f"Upload error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Dashboard Overview ──
@api_router.get("/dashboard/overview")
async def dashboard_overview(feedstock: str = "calcite", omega: int = 5):
    pipeline = [
        {"$match": {"feedstock": feedstock, "omega_threshold": omega, "cdr_t_yr": {"$ne": None, "$gt": 0}}},
        {"$group": {
            "_id": None,
            "total_cdr_t_yr":    {"$sum": "$cdr_t_yr"},
            "avg_cdr_t_yr":      {"$avg": "$cdr_t_yr"},
            "total_samples":     {"$sum": 1},
            "avg_ph":            {"$avg": "$ph"},
            "avg_alkalinity":    {"$avg": "$alkalinity"},
            "avg_rock_addition": {"$avg": "$rock_addition"},
            "avg_omega_final":   {"$avg": "$omega_final"},
        }}
    ]
    result = await db.erw_samples.aggregate(pipeline).to_list(1)
    total = await db.erw_samples.count_documents({"feedstock": feedstock, "omega_threshold": omega})
    successful = await db.erw_samples.count_documents({"feedstock": feedstock, "omega_threshold": omega, "success_flag": 1})
    r = result[0] if result else {}
    return {
        "total_cdr_t_yr":    r.get("total_cdr_t_yr", 0),
        "avg_cdr_t_yr":      r.get("avg_cdr_t_yr", 0),
        "total_samples":     total,
        "samples_with_cdr":  r.get("total_samples", 0),
        "avg_ph":            r.get("avg_ph", 0),
        "avg_alkalinity":    r.get("avg_alkalinity", 0),
        "avg_rock_addition": r.get("avg_rock_addition", 0),
        "avg_omega_final":   r.get("avg_omega_final", 0),
        "success_rate":      (successful / total * 100) if total > 0 else 0,
        "feedstock": feedstock, "omega_threshold": omega,
    }


# ── Full analytics data ──
@api_router.get("/analytics/full")
async def analytics_full(feedstock: str = "calcite", omega: int = 5):
    pipeline = [
        {"$match": {"feedstock": feedstock, "omega_threshold": omega}},
        {"$project": {
            "_id": 0, "ph": 1, "alkalinity": 1, "dic": 1, "pco2": 1, "fco2": 1,
            "temp_c": 1, "rock_addition": 1, "cdr_t_yr": 1, "omega_calcite": 1,
            "si_calcite": 1, "omega_final": 1, "discharge": 1, "region": 1, "state": 1,
            "ca": 1, "mg": 1, "na": 1, "k": 1, "hco3": 1, "co3": 1, "co2_aq": 1,
            "salinity": 1, "z_plus": 1, "z_minus": 1, "nicb": 1, "cl": 1, "so4": 1,
            "no3": 1, "latitude": 1, "longitude": 1, "river_name": 1, "sample_no": 1,
        }}
    ]
    return await db.erw_samples.aggregate(pipeline).to_list(2000)


# ── Basin stats ──
@api_router.get("/analytics/basin-stats")
async def basin_stats(feedstock: str = "calcite", omega: int = 5):
    pipeline = [
        {"$match": {"feedstock": feedstock, "omega_threshold": omega, "region": {"$ne": ""}}},
        {"$group": {
            "_id": "$region", "count": {"$sum": 1},
            "avg_ta": {"$avg": "$alkalinity"}, "avg_ca": {"$avg": "$ca"},
            "avg_mg": {"$avg": "$mg"}, "avg_na": {"$avg": "$na"}, "avg_k": {"$avg": "$k"},
            "avg_hco3": {"$avg": "$hco3"}, "avg_dic": {"$avg": "$dic"},
            "avg_pco2": {"$avg": "$pco2"}, "avg_co2_aq": {"$avg": "$co2_aq"},
            "avg_ph": {"$avg": "$ph"}, "avg_si_calcite": {"$avg": "$si_calcite"},
            "avg_omega_calcite": {"$avg": "$omega_calcite"},
            "total_cdr": {"$sum": "$cdr_t_yr"}, "avg_cdr": {"$avg": "$cdr_t_yr"},
            "avg_rock_add": {"$avg": "$rock_addition"},
        }},
        {"$sort": {"avg_ta": -1}}
    ]
    results = await db.erw_samples.aggregate(pipeline).to_list(50)
    return [{
        "basin": r["_id"], "count": r["count"],
        "avg_ta": r["avg_ta"] or 0, "avg_ca": r["avg_ca"] or 0, "avg_mg": r["avg_mg"] or 0,
        "avg_na": r["avg_na"] or 0, "avg_k": r["avg_k"] or 0,
        "avg_hco3": r["avg_hco3"] or 0, "avg_dic": r["avg_dic"] or 0,
        "avg_pco2": r["avg_pco2"] or 0, "avg_co2_aq": r["avg_co2_aq"] or 0,
        "avg_ph": r["avg_ph"] or 0, "avg_si_calcite": r["avg_si_calcite"] or 0,
        "avg_omega_calcite": r["avg_omega_calcite"] or 0,
        "total_cdr": r["total_cdr"] or 0, "avg_cdr": r["avg_cdr"] or 0,
        "avg_rock_add": r["avg_rock_add"] or 0,
        "ca_mg_ratio": ((r["avg_ca"] or 0) + (r["avg_mg"] or 0)) / max(((r["avg_na"] or 0) + (r["avg_k"] or 0)), 0.01),
    } for r in results]


# ── NICB quality ──
@api_router.get("/analytics/nicb-quality")
async def nicb_quality(feedstock: str = "calcite", omega: int = 5):
    pipeline = [
        {"$match": {"feedstock": feedstock, "omega_threshold": omega, "nicb": {"$ne": None}, "region": {"$ne": ""}}},
        {"$group": {
            "_id": "$region", "count": {"$sum": 1},
            "within_5":  {"$sum": {"$cond": [{"$lte": [{"$abs": "$nicb"}, 5]}, 1, 0]}},
            "within_10": {"$sum": {"$cond": [{"$and": [{"$gt": [{"$abs": "$nicb"}, 5]}, {"$lte": [{"$abs": "$nicb"}, 10]}]}, 1, 0]}},
            "beyond_10": {"$sum": {"$cond": [{"$gt": [{"$abs": "$nicb"}, 10]}, 1, 0]}},
        }},
        {"$sort": {"_id": 1}}
    ]
    results = await db.erw_samples.aggregate(pipeline).to_list(50)
    return [{
        "basin": r["_id"], "count": r["count"],
        "pct_within_5":  round(r["within_5"]  / max(r["count"], 1) * 100, 1),
        "pct_within_10": round(r["within_10"] / max(r["count"], 1) * 100, 1),
        "pct_beyond_10": round(r["beyond_10"] / max(r["count"], 1) * 100, 1),
    } for r in results]


# ── Summary ──
@api_router.get("/summary")
async def get_summary(feedstock: str = "calcite", omega: int = 5):
    return await db.summary_stats.find({"feedstock": feedstock, "omega_threshold": omega}, {"_id": 0}).to_list(100)


# ── Region CDR ──
@api_router.get("/regions/cdr")
async def regions_cdr(feedstock: str = "calcite", omega: int = 5):
    pipeline = [
        {"$match": {"feedstock": feedstock, "omega_threshold": omega, "region": {"$ne": ""}}},
        {"$group": {
            "_id": "$region", "total_cdr": {"$sum": "$cdr_t_yr"}, "avg_cdr": {"$avg": "$cdr_t_yr"},
            "count": {"$sum": 1}, "avg_ph": {"$avg": "$ph"}, "avg_rock_add": {"$avg": "$rock_addition"},
        }},
        {"$sort": {"total_cdr": -1}}
    ]
    results = await db.erw_samples.aggregate(pipeline).to_list(50)
    return [{"region": r["_id"], "total_cdr": r["total_cdr"] or 0, "avg_cdr": r["avg_cdr"] or 0,
             "count": r["count"], "avg_ph": r["avg_ph"] or 0, "avg_rock_add": r["avg_rock_add"] or 0} for r in results]


# ── Map data ──
@api_router.get("/samples/map")
async def get_map_data(feedstock: str = "calcite", omega: int = 5):
    pipeline = [
        {"$match": {"feedstock": feedstock, "omega_threshold": omega,
                    "latitude": {"$ne": None}, "longitude": {"$ne": None}}},
        {"$project": {"_id": 0, "latitude": 1, "longitude": 1, "river_name": 1,
                      "state": 1, "region": 1, "cdr_t_yr": 1, "alkalinity": 1,
                      "ph": 1, "rock_addition": 1, "omega_final": 1, "sample_no": 1,
                      "ca": 1, "mg": 1, "hco3": 1, "dic": 1, "si_calcite": 1}}
    ]
    return await db.erw_samples.aggregate(pipeline).to_list(2000)


# ── Samples ──
@api_router.get("/samples")
async def get_samples(feedstock: str = "calcite", omega: int = 5, region: Optional[str] = None,
                      state: Optional[str] = None, limit: int = 200, skip: int = 0):
    query = {"feedstock": feedstock, "omega_threshold": omega}
    if region: query["region"] = region
    if state:  query["state"] = state
    docs  = await db.erw_samples.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    total = await db.erw_samples.count_documents(query)
    return {"samples": docs, "total": total}


# ── Filters ──
@api_router.get("/filters")
async def get_filters(feedstock: str = "calcite", omega: int = 5):
    query = {"feedstock": feedstock, "omega_threshold": omega}
    regions = await db.erw_samples.distinct("region", query)
    states  = await db.erw_samples.distinct("state", query)
    return {"regions": sorted([r for r in regions if r]), "states": sorted([s for s in states if s])}


# ── Feedstocks ──
@api_router.get("/feedstocks")
async def list_feedstocks():
    return await db.feedstocks.find({}, {"_id": 0}).to_list(50)


# ── Comparison ──
@api_router.get("/comparison")
async def omega_comparison(feedstock: str = "calcite"):
    thresholds = await db.erw_samples.distinct("omega_threshold", {"feedstock": feedstock})
    results = []
    for omega in sorted(thresholds):
        pipeline = [
            {"$match": {"feedstock": feedstock, "omega_threshold": omega}},
            {"$group": {
                "_id": "$region",
                "total_cdr": {"$sum": "$cdr_t_yr"}, "avg_cdr": {"$avg": "$cdr_t_yr"},
                "avg_rock_add": {"$avg": "$rock_addition"}, "count": {"$sum": 1},
                "successful": {"$sum": {"$cond": [{"$eq": ["$success_flag", 1]}, 1, 0]}},
            }},
            {"$sort": {"total_cdr": -1}}
        ]
        regions = await db.erw_samples.aggregate(pipeline).to_list(50)
        total_pipeline = [
            {"$match": {"feedstock": feedstock, "omega_threshold": omega}},
            {"$group": {"_id": None, "total_cdr": {"$sum": "$cdr_t_yr"}, "avg_rock_add": {"$avg": "$rock_addition"}, "count": {"$sum": 1}}}
        ]
        totals = await db.erw_samples.aggregate(total_pipeline).to_list(1)
        ti = totals[0] if totals else {"total_cdr": 0, "avg_rock_add": 0, "count": 0}
        results.append({
            "omega_threshold": omega, "total_cdr": ti.get("total_cdr", 0) or 0,
            "avg_rock_add": ti.get("avg_rock_add", 0) or 0, "total_samples": ti.get("count", 0),
            "regions": [{"region": r["_id"], "total_cdr": r["total_cdr"] or 0, "avg_cdr": r["avg_cdr"] or 0,
                         "avg_rock_add": r["avg_rock_add"] or 0, "count": r["count"],
                         "success_rate": (r["successful"] / r["count"] * 100) if r["count"] > 0 else 0}
                        for r in regions if r["_id"]]
        })
    return results


# ── Top Rivers ──
@api_router.get("/rivers/top")
async def top_rivers(feedstock: str = "calcite", omega: int = 5, limit: int = 20):
    pipeline = [
        {"$match": {"feedstock": feedstock, "omega_threshold": omega, "river_name": {"$ne": ""}, "cdr_t_yr": {"$ne": None}}},
        {"$group": {
            "_id": "$river_name", "total_cdr": {"$sum": "$cdr_t_yr"}, "avg_cdr": {"$avg": "$cdr_t_yr"},
            "count": {"$sum": 1}, "region": {"$first": "$region"}, "state": {"$first": "$state"},
        }},
        {"$sort": {"total_cdr": -1}}, {"$limit": limit}
    ]
    results = await db.erw_samples.aggregate(pipeline).to_list(limit)
    return [{"river": r["_id"], "total_cdr": r["total_cdr"] or 0, "avg_cdr": r["avg_cdr"] or 0,
             "count": r["count"], "region": r.get("region", ""), "state": r.get("state", "")} for r in results]


# ── States CDR ──
@api_router.get("/states/cdr")
async def states_cdr(feedstock: str = "calcite", omega: int = 5):
    pipeline = [
        {"$match": {"feedstock": feedstock, "omega_threshold": omega, "state": {"$ne": ""}}},
        {"$group": {"_id": "$state", "total_cdr": {"$sum": "$cdr_t_yr"}, "avg_cdr": {"$avg": "$cdr_t_yr"}, "count": {"$sum": 1}}},
        {"$sort": {"total_cdr": -1}}
    ]
    results = await db.erw_samples.aggregate(pipeline).to_list(50)
    return [{"state": r["_id"], "total_cdr": r["total_cdr"] or 0, "avg_cdr": r["avg_cdr"] or 0, "count": r["count"]} for r in results]


# ── AI Chat ──
@api_router.post("/chat", response_model=ChatResponse)
async def chat_endpoint(req: ChatRequest):
    try:
        from openai import AsyncOpenAI
        openai_client = AsyncOpenAI(api_key=os.environ.get("OPENAI_API_KEY", ""))
        summary_docs   = await db.summary_stats.find({}, {"_id": 0}).to_list(100)
        feedstock_docs = await db.feedstocks.find({}, {"_id": 0}).to_list(50)
        total   = await db.erw_samples.count_documents({})
        regions = await db.erw_samples.distinct("region")
        states  = await db.erw_samples.distinct("state")
        data_context = f"""You are an expert in Enhanced Rock Weathering (ERW) and carbon dioxide removal (CDR) for Indian rivers.
Available data: {total} samples across regions: {', '.join([r for r in regions if r])}
States: {', '.join([s for s in states if s])}
Feedstocks: {json.dumps([d.get('name') for d in feedstock_docs])}
Summary stats: {json.dumps(summary_docs, indent=2, default=str)}
Key: ERW adds crushed minerals to rivers to capture CO2. CDR in t CO2/yr. Omega = calcite saturation. Rock addition in mol/kg.
Answer accurately using data. If unsure, say so."""
        response = await openai_client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "system", "content": data_context}, {"role": "user", "content": req.message}],
            max_tokens=1000,
        )
        reply = response.choices[0].message.content
        now = datetime.now(timezone.utc).isoformat()
        await db.chat_messages.insert_many([
            {"session_id": req.session_id, "role": "user",      "content": req.message, "timestamp": now},
            {"session_id": req.session_id, "role": "assistant", "content": reply,        "timestamp": now},
        ])
        return ChatResponse(reply=reply, session_id=req.session_id)
    except Exception as e:
        logger.error(f"Chat error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Simulator ──
@api_router.get("/simulator/full-data")
async def simulator_full_data():
    pipeline = [{"$project": {
        "_id": 0, "feedstock": 1, "omega_threshold": 1, "data_category": 1,
        "latitude": 1, "longitude": 1, "region": 1, "state": 1,
        "river_name": 1, "river_type": 1, "sample_no": 1,
        "ph": 1, "alkalinity": 1, "ca": 1, "mg": 1, "pco2": 1,
        "si_calcite": 1, "omega_calcite": 1, "rock_addition": 1,
        "cdr_t_yr": 1, "cdr_kt_yr": 1, "discharge": 1,
        "ph_final": 1, "alk_final": 1, "ca_final": 1,
        "omega_final": 1, "pco2_final": 1, "dic_final": 1, "success_flag": 1,
    }}]
    return await db.erw_samples.aggregate(pipeline).to_list(5000)


@api_router.get("/simulator/available-combos")
async def simulator_combos():
    pipeline = [
        {"$group": {
            "_id": {"feedstock": "$feedstock", "omega": "$omega_threshold", "category": "$data_category"},
            "count": {"$sum": 1}, "total_cdr": {"$sum": "$cdr_t_yr"}, "avg_ph": {"$avg": "$ph"},
        }},
        {"$sort": {"_id.feedstock": 1, "_id.omega": 1}}
    ]
    results = await db.erw_samples.aggregate(pipeline).to_list(50)
    return [{
        "feedstock": r["_id"]["feedstock"], "omega": r["_id"]["omega"],
        "data_category": r["_id"].get("category", "baseline"),
        "count": r["count"], "total_cdr": r.get("total_cdr") or 0, "avg_ph": r.get("avg_ph") or 0,
    } for r in results]


@api_router.get("/chat/history")
async def chat_history(session_id: str = "default", limit: int = 50):
    return await db.chat_messages.find({"session_id": session_id}, {"_id": 0}).sort("timestamp", 1).to_list(limit)


app.include_router(api_router)
app.add_middleware(
    CORSMiddleware, allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"], allow_headers=["*"],
)
